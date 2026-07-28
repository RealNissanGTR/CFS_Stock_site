import hashlib
import hmac
import io
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Optional
from fastapi import FastAPI, Form, Request, status, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook

from db.db_init import (
    SessionLocal,
    User,
    Stock,
    Logs,
    init_db,
    verify_password,
    add_user_full,
    create_stock_item,
    delete_stock_item,
    add_inventory_item,
    remove_inventory_item,
    move_inventory_item,
    get_user_by_id,
    update_user,
    delete_user,
)

from backup.backup_db import quick_backup

BASE_DIR = Path(__file__).resolve().parent
FRONTEND_DIR = BASE_DIR.parent / "FrontEnd"

app = FastAPI(title="CFS Stock")
app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")
templates = Jinja2Templates(directory=str(FRONTEND_DIR))

SECRET_KEY = os.getenv("SECRET_KEY", "change-this-secret")
COOKIE_NAME = "cfs_auth"
COOKIE_MAX_AGE = 3600

init_db()


def _sign_payload(payload: str) -> str:
    return hmac.new(SECRET_KEY.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()


def _make_cookie_value(user: User) -> str:
    payload = f"{user.username}:{int(user.is_admin)}"
    return f"{payload}:{_sign_payload(payload)}"


def _parse_cookie_value(cookie_value: Optional[str]) -> Optional[tuple[str, bool]]:
    if not cookie_value:
        return None

    parts = cookie_value.rsplit(":", 1)
    if len(parts) != 2:
        return None

    payload, signature = parts
    if not hmac.compare_digest(signature, _sign_payload(payload)):
        return None

    payload_parts = payload.rsplit(":", 1)
    if len(payload_parts) != 2:
        return None

    username, is_admin_text = payload_parts
    return username, bool(int(is_admin_text))


def get_current_user(request: Request) -> Optional[User]:
    cookie_value = request.cookies.get(COOKIE_NAME)
    parsed = _parse_cookie_value(cookie_value)
    if not parsed:
        return None

    username, _ = parsed
    with SessionLocal() as db:
        return db.query(User).filter_by(username=username).first()


def require_login(request: Request):
    user = get_current_user(request)
    if user is None:
        raise HTTPException(
            status_code=status.HTTP_303_SEE_OTHER,
            headers={"Location": "/login"},
        )
    return user


def require_admin(request: Request):
    user = require_login(request)
    if not user.is_admin:
        raise HTTPException(
            status_code=status.HTTP_303_SEE_OTHER,
            headers={"Location": "/home"},
        )
    return user


@app.get("/", response_class=HTMLResponse)
def root(request: Request):
    return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request, "error": None})


@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...)):
    with SessionLocal() as db:
        user = db.query(User).filter_by(username=username.strip()).first()
        if user and verify_password(password.strip(), user.password_hash):
            response = RedirectResponse(url="/home", status_code=status.HTTP_303_SEE_OTHER)
            response.set_cookie(
                key=COOKIE_NAME,
                value=_make_cookie_value(user),
                httponly=True,
                max_age=COOKIE_MAX_AGE,
                expires=COOKIE_MAX_AGE,
                samesite="lax",
            )
            return response

    return templates.TemplateResponse(
        "login.html",
        {"request": request, "error": "Invalid username or password."},
        status_code=status.HTTP_401_UNAUTHORIZED,
    )


@app.get("/logout")
def logout():
    response = RedirectResponse(url="/", status_code=status.HTTP_303_SEE_OTHER)
    response.delete_cookie(COOKIE_NAME)
    return response


@app.get("/overview", response_class=HTMLResponse)
def overview(request: Request):
    user = require_login(request)
    with SessionLocal() as db:
        stocks = db.query(Stock).order_by(
            Stock.category,
            Stock.product_name,
            Stock.item_code,
            Stock.location,
        ).all()

    grouped = {}
    for stock in stocks:
        key = (stock.item_code, stock.product_name, stock.category or "Unassigned")
        if key not in grouped:
            grouped[key] = {
                "category": stock.category or "Unassigned",
                "item_code": stock.item_code,
                "product_name": stock.product_name,
                "workshop": 0,
                "paddock": 0,
                "other": [],
                "total": 0,
            }

        count = stock.stock_count or 0
        location_key = stock.location.strip().lower()
        if location_key == "workshop":
            grouped[key]["workshop"] += count
        elif location_key == "paddock":
            grouped[key]["paddock"] += count
        else:
            grouped[key]["other"].append(f"{stock.location}: {count}")

        grouped[key]["total"] += count

    return templates.TemplateResponse(
        "overview.html",
        {
            "request": request,
            "user": user,
            "products": list(grouped.values()),
        },
    )


@app.get("/home", response_class=HTMLResponse)
def home(request: Request):
    user = require_login(request)
    flash = get_flash(request)
    stock_form_state = get_stock_form_state(request)

    with SessionLocal() as db:
        stocks = db.query(Stock).order_by(
            Stock.category,
            Stock.product_name,
            Stock.item_code,
            Stock.location,
        ).all()

        categories = sorted({s.category or "Unassigned" for s in stocks})
        products_map = {}
        for s in stocks:
            key = s.item_code
            if key not in products_map:
                products_map[key] = {
                    "item_code": s.item_code,
                    "product_name": s.product_name,
                    "category": s.category or "Unassigned",
                    "workshop": 0,
                    "paddock": 0,
                    "total": 0,
                }
            count = s.stock_count or 0
            loc = (s.location or "").strip().lower()
            if loc == "workshop":
                products_map[key]["workshop"] += count
            elif loc == "paddock":
                products_map[key]["paddock"] += count
            products_map[key]["total"] += count

    response = templates.TemplateResponse("home.html", {
        "request": request,
        "user": user,
        "categories": categories,
        "products": list(products_map.values()),
        "selected_item_code": request.query_params.get("item_code", "").strip() or stock_form_state.get("item_code", ""),
        "stock_form_state": stock_form_state,
        "message": flash,
    })
    response.delete_cookie(FLASH_COOKIE)
    return response


@app.get("/admin", response_class=HTMLResponse)
def admin_panel(request: Request):
    try:
        user = require_admin(request)
    except Exception:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)

    flash = get_flash(request)

    try:
        with SessionLocal() as db:
            stocks_raw = db.query(Stock).order_by(
                Stock.category, Stock.product_name, Stock.item_code, Stock.location,
            ).all()

            stocks_dict = [
                {
                    "id": s.id,
                    "item_code": s.item_code or "",
                    "product_name": s.product_name or "",
                    "category": s.category or "",
                    "stock_count": s.stock_count or 0,
                    "location": s.location or "",
                }
                for s in stocks_raw
            ]

            users_list = [
                {"id": u.id, "username": u.username, "is_admin": u.is_admin}
                for u in db.query(User).order_by(User.username).all()
            ]

            logs_list = [
                {
                    "timestamp": l.timestamp,
                    "action": l.action,
                    "stock_name": l.stock_name,
                    "category": l.category or "",
                    "quantity": l.quantity,
                    "location": l.location,
                    "performed_by": l.performed_by or "",
                    "work_order": l.work_order or "",
                    "purchase_order": l.purchase_order or "",
                    "comments": l.comments or "",
                }
                for l in db.query(Logs).order_by(Logs.timestamp.desc()).limit(200).all()
            ]

        response = templates.TemplateResponse("admin.html", {
            "request": request,
            "user": user,
            "stocks": stocks_dict,
            "users": users_list,
            "logs": logs_list,
            "message": flash,
        })
        response.delete_cookie(FLASH_COOKIE)
        return response

    except Exception as e:
        print(f"Admin panel error: {e}")
        r = RedirectResponse(url="/home", status_code=status.HTTP_303_SEE_OTHER)
        return set_flash(r, f"Admin panel error: {str(e)}")


def _normalize_item_fields(
    item_code: Optional[str],
    product_name: Optional[str],
    stock_name: Optional[str] = None,
):
    normalized_item_code = (item_code or stock_name or "").strip()
    normalized_product_name = (product_name or stock_name or "").strip()
    return normalized_item_code, normalized_product_name


def _normalize_item_code(item_code: Optional[str], stock_name: Optional[str] = None):
    return (item_code or stock_name or "").strip()


@app.post("/admin/create_stock")
def admin_create_stock(
    request: Request,
    item_code: str = Form(...),
    product_name: str = Form(...),
    quantity: int = Form(...),
    location: str = Form("Workshop"),
    category: str = Form(""),
):
    try:
        user = require_admin(request)
        
        item_code = item_code.strip()
        product_name = product_name.strip()
        
        if not item_code or not product_name:
            r = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
            return set_flash(r, "Item code and product name are required.")
        
        success = create_stock_item(
            item_code=item_code,
            product_name=product_name,
            quantity=quantity,
            location=location.strip(),
            category=category.strip() or None,
            performed_by=user.username,
            created_by_admin=True,
        )
        
        if success:
            quick_backup()
            message = f"Stock item '{product_name}' created successfully."
        else:
            message = f"Stock item with code '{item_code}' already exists at {location}."
        
        r = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
        return set_flash(r, message)
        
    except Exception as e:
        r = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
        return set_flash(r, f"Error creating stock: {str(e)}")


@app.post("/admin/add_user")
def admin_add_user(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    is_admin: str = Form(""),
):
    require_admin(request)
    success = add_user_full(
        username=username.strip(),
        password=password.strip(),
        is_admin=bool(is_admin),
    )
    message = f"User '{username}' created." if success else f"Username '{username}' already exists."
    r = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
    return set_flash(r, message)


@app.post("/admin/toggle_admin/{user_id}")
def admin_toggle_admin(request: Request, user_id: int):
    require_admin(request)
    with SessionLocal() as db:
        target = db.query(User).filter_by(id=user_id).first()
        if target:
            target.is_admin = not target.is_admin
            db.commit()
            message = f"{'Admin granted to' if target.is_admin else 'Admin revoked from'} '{target.username}'."
        else:
            message = "User not found."
    r = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
    return set_flash(r, message)


@app.post("/admin/edit_user")
def admin_edit_user(
    request: Request,
    user_id: int = Form(...),
    new_username: str = Form(""),
    new_password: str = Form(""),
    is_admin: str = Form(None),
):
    admin = require_admin(request)
    success = update_user(
        user_id=user_id,
        new_username=new_username.strip() or None,
        new_password=new_password.strip() or None,
        is_admin=bool(is_admin),
    )
    message = "User updated successfully." if success else "Unable to update user."
    response = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
    set_flash(response, message)

    if success and user_id == admin.id:
        updated = get_user_by_id(user_id)
        if updated:
            response.set_cookie(
                key=COOKIE_NAME, value=_make_cookie_value(updated),
                httponly=True, max_age=COOKIE_MAX_AGE,
                expires=COOKIE_MAX_AGE, samesite="lax",
            )
    return response


@app.post("/admin/delete_user/{user_id}")
def admin_delete_user(request: Request, user_id: int):
    try:
        admin = require_admin(request)
        success = delete_user(user_id=user_id, protect_username=admin.username)
        message = "User deleted successfully." if success else "Unable to delete user (may be the last admin or yourself)."
        r = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
        return set_flash(r, message)
    except Exception as e:
        r = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
        return set_flash(r, f"Error deleting user: {str(e)}")


@app.post("/inventory/change")
def inventory_change(
    request: Request,
    item_code: str = Form(...),
    location: str = Form(...),
    quantity: int = Form(...),
    direction: str = Form(...),
    order_code: str = Form(...),
    comments: str = Form(""),
    category: str = Form(""),
):
    try:
        user = require_login(request)
        item_code = item_code.strip()
        location = location.strip()
        direction = direction.strip().lower()
        order_code = order_code.strip()
        comments = comments.strip()
        category = category.strip()
        stock_form_state = {
            "category": category,
            "item_code": item_code,
            "location": location,
            "quantity": str(quantity),
            "direction": direction if direction in {"add", "remove"} else "add",
            "order_code": order_code,
            "comments": comments,
        }

        if not item_code or quantity <= 0 or not order_code or direction not in {"add", "remove"}:
            r = RedirectResponse(url="/home", status_code=status.HTTP_303_SEE_OTHER)
            set_flash(r, "Invalid input. Please fill all required fields.")
            return set_stock_form_state(r, stock_form_state)

        if direction == "add":
            success = add_inventory_item(
                item_code=item_code, quantity=quantity, location=location,
                category=category or None, performed_by=user.username,
                work_order=order_code, comments=comments or None,
                user_is_admin=user.is_admin,
            )
            message = "Inventory added successfully." if success else "Unable to add inventory."
        else:
            success = remove_inventory_item(
                item_code=item_code, quantity=quantity, location=location,
                category=category or None, performed_by=user.username,
                purchase_order=order_code, comments=comments or None,
            )
            message = "Inventory removed successfully." if success else "Unable to remove inventory."

        quick_backup()
        r = RedirectResponse(url="/home", status_code=status.HTTP_303_SEE_OTHER)
        set_flash(r, message)
        return set_stock_form_state(r, stock_form_state)

    except Exception as e:
        r = RedirectResponse(url="/home", status_code=status.HTTP_303_SEE_OTHER)
        set_flash(r, f"Error: {str(e)}")
        return set_stock_form_state(r, {
            "category": category.strip(),
            "item_code": item_code.strip(),
            "location": location.strip(),
            "quantity": str(quantity),
            "direction": direction.strip().lower() if direction else "add",
            "order_code": order_code.strip(),
            "comments": comments.strip(),
        })


@app.post("/inventory/move")
def inventory_move(
    request: Request,
    item_code: str = Form(...),
    from_location: str = Form(...),
    to_location: str = Form(...),
    quantity: int = Form(...),
    comments: str = Form(""),
    category: str = Form(""),
):
    try:
        user = require_login(request)
        success = move_inventory_item(
            item_code=item_code.strip(), quantity=quantity,
            from_location=from_location.strip(), to_location=to_location.strip(),
            performed_by=user.username, comments=comments.strip() or None,
        )
        quick_backup()
        message = "Stock moved successfully." if success else "Unable to move stock. Check quantity and locations."
        r = RedirectResponse(url="/home", status_code=status.HTTP_303_SEE_OTHER)
        return set_flash(r, message)
    except Exception as e:
        r = RedirectResponse(url="/home", status_code=status.HTTP_303_SEE_OTHER)
        return set_flash(r, f"Error moving stock: {str(e)}")


@app.post("/admin/edit_stock")
def admin_edit_stock(
    request: Request,
    stock_id: int = Form(...),
    new_item_code: str = Form(""),
    new_product_name: str = Form(""),
    new_category: str = Form(""),
    new_quantity: int = Form(None),
):
    try:
        user = require_login(request)
        if not user.is_admin:
            r = RedirectResponse(url="/home", status_code=status.HTTP_303_SEE_OTHER)
            return set_flash(r, "Admin only.")

        with SessionLocal() as db:
            stock = db.query(Stock).filter(Stock.id == stock_id).first()
            if not stock:
                message = "Stock item not found."
            else:
                old_name = stock.product_name
                if new_item_code.strip():
                    stock.item_code = new_item_code.strip()
                if new_product_name.strip():
                    stock.product_name = new_product_name.strip()
                if new_category.strip():
                    stock.category = new_category.strip()
                if new_quantity is not None and new_quantity >= 0:
                    stock.stock_count = new_quantity
                db.commit()
                quick_backup()
                message = f"Stock item '{old_name}' updated successfully."

        r = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
        return set_flash(r, message)
    except Exception as e:
        r = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
        return set_flash(r, f"Error updating stock: {str(e)}")


@app.post("/inventory/delete")
def inventory_delete(
    request: Request,
    item_code: str = Form(...),
    location: str = Form(...),
):
    try:
        user = require_login(request)
        if not user.is_admin:
            r = RedirectResponse(url="/home", status_code=status.HTTP_303_SEE_OTHER)
            return set_flash(r, "Admin only.")

        item_code = item_code.strip()
        location = location.strip()

        if not item_code or not location:
            r = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
            return set_flash(r, "Item code and location are required.")

        # Use delete_stock_item so the action is logged
        success = delete_stock_item(
            item_code=item_code,
            location=location,
            performed_by=user.username,
            deleted_by_admin=True,
        )

        if success:
            quick_backup()
            message = f"Stock item '{item_code}' deleted successfully."
        else:
            message = f"Stock item '{item_code}' not found at {location}."

        r = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
        return set_flash(r, message)

    except Exception as e:
        r = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
        return set_flash(r, f"Error deleting stock: {str(e)}")


@app.get("/admin/export_data")
def admin_export_data(request: Request):
    require_admin(request)

    with SessionLocal() as db:
        stocks = db.query(Stock).order_by(Stock.item_code, Stock.location).all()
        logs = db.query(Logs).order_by(Logs.timestamp.desc()).all()
        users = db.query(User).order_by(User.username).all()

    workbook = Workbook()

    stock_sheet = workbook.active
    stock_sheet.title = "Stock"
    stock_sheet.append(["Item Code", "Product Name", "Category", "Location", "Stock Count"])
    for stock in stocks:
        stock_sheet.append([
            stock.item_code,
            stock.product_name,
            stock.category or "",
            stock.location,
            stock.stock_count,
        ])

    logs_sheet = workbook.create_sheet("Logs")
    logs_sheet.append([
        "Timestamp",
        "Action",
        "Stock Name",
        "Category",
        "Quantity",
        "Location",
        "Performed By",
        "Work Order",
        "Purchase Order",
        "Comments",
    ])
    for log in logs:
        logs_sheet.append([
            log.timestamp,
            log.action,
            log.stock_name,
            log.category or "",
            log.quantity,
            log.location,
            log.performed_by or "",
            log.work_order or "",
            log.purchase_order or "",
            log.comments or "",
        ])

    users_sheet = workbook.create_sheet("Users")
    users_sheet.append(["ID", "Username", "Is Admin"])
    for user in users:
        users_sheet.append([user.id, user.username, bool(user.is_admin)])

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"data-export-{timestamp}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    if exc.status_code == 303:
        response = RedirectResponse(url=exc.headers.get("location", "/home"), status_code=exc.status_code)
        return response
    
    message = "An error occurred. Please try again."
    return RedirectResponse(
        url=f"/home?message={message}",
        status_code=status.HTTP_303_SEE_OTHER,
    )


@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc: Exception):
    message = "An unexpected error occurred. Please try again."
    return RedirectResponse(
        url=f"/home?message={message}",
        status_code=status.HTTP_303_SEE_OTHER,
    )


FLASH_COOKIE = "cfs_flash"
STOCK_FORM_COOKIE = "cfs_stock_form"

def set_flash(response: RedirectResponse, message: str) -> RedirectResponse:
    response.set_cookie(FLASH_COOKIE, message, max_age=10, httponly=True, samesite="lax")
    return response

def get_flash(request: Request) -> str:
    return request.cookies.get(FLASH_COOKIE, "")

def set_stock_form_state(response: RedirectResponse, form_state: dict[str, str]) -> RedirectResponse:
    payload = {
        "category": (form_state.get("category") or "").strip(),
        "item_code": (form_state.get("item_code") or "").strip(),
        "location": (form_state.get("location") or "").strip(),
        "quantity": str(form_state.get("quantity") or "1").strip(),
        "direction": (form_state.get("direction") or "add").strip(),
        "order_code": (form_state.get("order_code") or "").strip(),
        "comments": (form_state.get("comments") or "").strip(),
    }
    response.set_cookie(
        STOCK_FORM_COOKIE,
        json.dumps(payload, separators=(",", ":")),
        max_age=86400,
        httponly=True,
        samesite="lax",
    )
    return response

def get_stock_form_state(request: Request) -> dict[str, str]:
    raw = request.cookies.get(STOCK_FORM_COOKIE)
    if not raw:
        return {}

    try:
        data = json.loads(raw)
    except Exception:
        return {}

    if not isinstance(data, dict):
        return {}

    return {
        "category": str(data.get("category") or ""),
        "item_code": str(data.get("item_code") or ""),
        "location": str(data.get("location") or ""),
        "quantity": str(data.get("quantity") or "1"),
        "direction": str(data.get("direction") or "add"),
        "order_code": str(data.get("order_code") or ""),
        "comments": str(data.get("comments") or ""),
    }