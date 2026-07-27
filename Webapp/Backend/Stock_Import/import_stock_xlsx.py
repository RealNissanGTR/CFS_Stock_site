from pathlib import Path
import sys

BASE_DIR = Path(__file__).resolve().parent
BACKEND_DIR = BASE_DIR.parent
sys.path.insert(0, str(BACKEND_DIR))

from openpyxl import load_workbook
from db.db_init import SessionLocal, Stock

def import_xlsx(path: Path, default_location: str = "Workshop"):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_labels = [str(cell).strip().lower() if cell else "" for cell in first_row]
    has_header = any(
        label in ("category", "catagory", "product code", "item code", "product name", "name")
        for label in header_labels
    )
    start_row = 2 if has_header else 1

    last_category = None
    with SessionLocal() as db:
        for row in ws.iter_rows(min_row=start_row, max_col=3, values_only=True):
            raw_category, item_code, product_name = row

            if raw_category is not None and str(raw_category).strip():
                last_category = str(raw_category).strip()

            category = last_category
            item_code = str(item_code).strip() if item_code else ""
            product_name = str(product_name).strip() if product_name else ""

            if not item_code or not product_name:
                continue

            stock = db.query(Stock).filter_by(item_code=item_code, location=default_location).first()
            if stock:
                stock.product_name = product_name
                stock.category = category
            else:
                db.add(Stock(
                    item_code=item_code,
                    product_name=product_name,
                    category=category,
                    stock_count=0,
                    location=default_location,
                ))

        db.commit()

if __name__ == "__main__":
    import_xlsx(BASE_DIR / "FILE_DIR_HERE.xlsx")